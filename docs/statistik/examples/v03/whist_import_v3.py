#!/usr/bin/env python3
"""Generate the first app-shaped v3 historical Whist data bundle."""

from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import whist_import_audit_v3 as audit_v3


ROOT = Path(__file__).resolve().parent
PLAYERS = audit_v3.PLAYERS
PRIMARY_SHEET = audit_v3.PRIMARY_SHEET
AUDIT_SHEET = audit_v3.AUDIT_SHEET


def slug(value: str) -> str:
    return (
        value.lower()
        .replace("å", "aa")
        .replace("æ", "ae")
        .replace("ø", "oe")
        .replace(" ", "_")
    )


def clean_text(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    return text or None


def normalize_game_marker(value: Any) -> int | None:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return None


def parse_bid_tricks(value: Any) -> int | None:
    text = clean_text(value)
    if not text:
        return None
    match = re.search(r"\b(7|8|9|10|11|12|13)\b", text)
    return int(match.group(1)) if match else None


def normalize_game_type(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    lowered = text.lower()
    if "ren sol" in lowered:
        return "ren_sol"
    if "sol" in lowered:
        return "sol"
    if "vip" in lowered:
        return "vip"
    if "halv" in lowered:
        return "halve"
    if "sans" in lowered:
        return "sans"
    if "gode" in lowered:
        return "gode"
    if "alm" in lowered:
        return "alm"
    return slug(text)


def session_id(session_number: str, date: str | None) -> str:
    date_part = date or "unknown-date"
    return f"session_{slug(session_number)}_{date_part}"


def game_id_for(session: dict[str, Any], sequence: int) -> str:
    return f"{session['id']}_game_{sequence:03d}"


def control_sessions(workbook: Any) -> dict[str, dict[str, Any]]:
    ws = workbook[AUDIT_SHEET]
    sessions: dict[str, dict[str, Any]] = {}
    for row in range(3, ws.max_row + 1):
        session_number = audit_v3.normalize_session(ws.cell(row, 1).value)
        if not audit_v3.is_session_key(session_number):
            continue
        sessions[session_number] = {
            "sessionNumber": session_number,
            "date": audit_v3.normalize_date(ws.cell(row, 3).value),
            "location": clean_text(ws.cell(row, 12).value),
            "expectedGameCount": audit_v3.normalize_int(ws.cell(row, 9).value),
            "controlRow": row,
        }
    return sessions


def base_sessions(workbook: Any, audit_result: dict[str, Any]) -> dict[str, dict[str, Any]]:
    controls = control_sessions(workbook)
    validation = {
        row["session"]: row for row in audit_result.get("sessions", [])
    }
    sessions: dict[str, dict[str, Any]] = {}
    for session_number, control in controls.items():
        validation_row = validation.get(session_number, {})
        sid = session_id(session_number, control["date"])
        sessions[session_number] = {
            "id": sid,
            "sessionNumber": session_number,
            "date": control["date"],
            "location": control["location"],
            "sourceSheetName": validation_row.get("importedCountSource") or AUDIT_SHEET,
            "expectedGameCount": control["expectedGameCount"],
            "importedGameCount": 0,
            "missingScoreRows": 0,
            "qualityStatus": validation_row.get("status", "pending"),
            "cumulativeBlockStartColumn": None,
            "deltaBlockStartColumn": None,
            "preferredScoreBlockNumericRows": None,
            "headerRow": None,
            "columnMapping": None,
        }
    return sessions


def scores_from_values(values: dict[str, Any]) -> dict[str, int] | None:
    scores: dict[str, int] = {}
    for player in PLAYERS:
        value = values.get(player)
        if not isinstance(value, (int, float)):
            return None
        scores[player] = int(value)
    return scores


def score_values_by_header(
    ws: Any,
    row: int,
    start_col: int,
    end_col: int,
    header_row: int,
) -> dict[str, Any] | None:
    values: dict[str, Any] = {}
    for col in range(start_col, end_col + 1):
        player = audit_v3.normalize_player(ws.cell(header_row, col).value)
        if player is None:
            return None
        values[player] = ws.cell(row, col).value
    return values


def game_quality_flags(
    scores: dict[str, int],
    game_type: str | None,
    bidder_id: str | None,
    dealer_id: str | None,
    partner_id: str | None,
    manual_review_required: bool = False,
    limited_source: bool = False,
) -> list[str]:
    flags: list[str] = []
    if sum(scores.values()) != 0:
        flags.append("score_sum_not_zero")
    if game_type is None:
        flags.append("missing_game_type")
    if bidder_id is None:
        flags.append("missing_bidder_or_winner")
    if dealer_id is None:
        flags.append("missing_dealer")
    if partner_id is None:
        flags.append("missing_partner")
    if manual_review_required:
        flags.append("manual_review_required")
    if limited_source:
        flags.append("limited_source")
    return flags


def append_game(
    session: dict[str, Any],
    sequence: int,
    source_game_marker: int | None,
    scores: dict[str, int],
    games: list[dict[str, Any]],
    player_results: list[dict[str, Any]],
    source_sheet_name: str,
    source_row: int,
    score_source: str,
    game_type_raw: str | None = None,
    bidder_id: str | None = None,
    partner_id: str | None = None,
    dealer_id: str | None = None,
    manual_review_required: bool = False,
    limited_source: bool = False,
) -> None:
    game_type_normalized = normalize_game_type(game_type_raw)
    game_id = game_id_for(session, sequence)
    flags = game_quality_flags(
        scores,
        game_type_normalized,
        bidder_id,
        dealer_id,
        partner_id,
        manual_review_required=manual_review_required,
        limited_source=limited_source,
    )
    checksum = sum(scores.values())
    winner_ids = [player for player in PLAYERS if scores[player] > 0]
    game = {
        "id": game_id,
        "sessionId": session["id"],
        "sessionNumber": session["sessionNumber"],
        "gameNumberInSession": sequence,
        "sourceGameMarker": source_game_marker,
        "gameTypeRaw": game_type_raw,
        "gameTypeNormalized": game_type_normalized,
        "bidTricks": parse_bid_tricks(game_type_raw),
        "bidderId": bidder_id,
        "bidderIds": [bidder_id] if bidder_id else [],
        "winnerId": winner_ids[0] if len(winner_ids) == 1 else None,
        "winnerIds": winner_ids,
        "partnerId": partner_id,
        "dealerId": dealer_id,
        "checksum": checksum,
        "scoreSource": score_source,
        "sourceSheetName": source_sheet_name,
        "sourceRow": source_row,
        "qualityFlags": flags,
    }
    games.append(game)
    for player in PLAYERS:
        player_results.append(
            {
                "id": f"{game_id}_{player}",
                "gameId": game_id,
                "playerId": player,
                "score": scores[player],
                "sourceSheetName": source_sheet_name,
                "sourceRow": source_row,
            }
        )


def import_individual_sheet_games(
    workbook: Any,
    sessions: dict[str, dict[str, Any]],
    games: list[dict[str, Any]],
    player_results: list[dict[str, Any]],
) -> None:
    for session_number, config in audit_v3.INDIVIDUAL_SESSION_SHEETS.items():
        if config["status"] == "empty_source_sheet":
            continue
        session = sessions[session_number]
        ws = workbook[config["sheet"]]
        max_data_row = config.get("maxDataRow", ws.max_row)
        delta_columns = config["deltaColumns"]
        cumulative_columns = config["cumulativeColumns"]
        header_row = 3 if session_number in {"1", "3", "4"} else 2
        previous_cumulative: dict[str, int] | None = None
        sequence = 0
        for row in range(1, max_data_row + 1):
            marker = normalize_game_marker(ws.cell(row, 1).value)
            if marker is None or row == 1:
                continue
            scores: dict[str, int] | None = None
            score_source = "delta_columns"
            if delta_columns is not None:
                delta_values = score_values_by_header(
                    ws, row, delta_columns[0], delta_columns[1], header_row
                )
                scores = scores_from_values(delta_values)
                if scores is None:
                    cumulative_values = score_values_by_header(
                        ws, row, cumulative_columns[0], cumulative_columns[1], header_row
                    )
                    scores = scores_from_values(cumulative_values)
                    score_source = "cumulative_columns"
            else:
                current_values = score_values_by_header(
                    ws, row, cumulative_columns[0], cumulative_columns[1], header_row
                )
                current = scores_from_values(current_values)
                if current is None:
                    continue
                if previous_cumulative is None:
                    scores = current
                else:
                    scores = {
                        player: current[player] - previous_cumulative[player]
                        for player in PLAYERS
                    }
                previous_cumulative = current
                score_source = "cumulative_diff"
            if scores is None:
                continue
            sequence += 1
            append_game(
                session,
                sequence,
                marker,
                scores,
                games,
                player_results,
                config["sheet"],
                row,
                score_source,
                limited_source=True,
            )
        session["importedGameCount"] = sequence
        session["sourceSheetName"] = config["sheet"]
        session["qualityStatus"] = config["status"]
        session["missingScoreRows"] = 0
        session["cumulativeBlockStartColumn"] = (
            cumulative_columns[0] if cumulative_columns else None
        )
        session["deltaBlockStartColumn"] = delta_columns[0] if delta_columns else None
        session["preferredScoreBlockNumericRows"] = sequence
        session["headerRow"] = header_row


def compound_bidder_partner(raw: Any) -> tuple[str | None, str | None]:
    result = audit_v3.normalize_compound_bidder_partner(raw)
    if result is None:
        return None, None
    return result


def import_primary_sheet_games(
    workbook: Any,
    sessions: dict[str, dict[str, Any]],
    games: list[dict[str, Any]],
    player_results: list[dict[str, Any]],
) -> None:
    ws = workbook[PRIMARY_SHEET]
    sequence_by_session: defaultdict[str, int] = defaultdict(int)
    for row in range(4, ws.max_row + 1):
        session_number = audit_v3.normalize_session(ws.cell(row, 1).value)
        marker = normalize_game_marker(ws.cell(row, 2).value)
        if session_number is None or marker is None:
            continue
        if session_number in audit_v3.INDIVIDUAL_SESSION_SHEETS:
            continue
        session = sessions.get(session_number)
        if session is None:
            continue
        scores = scores_from_values(
            {
                "Thomas": ws.cell(row, 21).value,
                "Peter": ws.cell(row, 22).value,
                "Janus": ws.cell(row, 23).value,
                "Christian": ws.cell(row, 24).value,
            }
        )
        if scores is None:
            continue
        bidder_raw = ws.cell(row, 8).value
        partner_raw = ws.cell(row, 9).value
        dealer_raw = ws.cell(row, 10).value
        bidder_id, partner_id = compound_bidder_partner(bidder_raw)
        if bidder_id is None:
            if audit_v3.is_self_partner_marker(bidder_raw):
                bidder_id = audit_v3.normalize_player(dealer_raw)
                partner_id = None
            else:
                bidder_id = audit_v3.normalize_player(bidder_raw)
        if partner_id is None and not audit_v3.is_self_partner_marker(partner_raw):
            partner_id = audit_v3.normalize_player(partner_raw)
        dealer_id = audit_v3.normalize_player(dealer_raw)
        sequence_by_session[session_number] += 1
        append_game(
            session,
            sequence_by_session[session_number],
            marker,
            scores,
            games,
            player_results,
            PRIMARY_SHEET,
            row,
            "primary_delta_columns",
            game_type_raw=clean_text(ws.cell(row, 11).value) or clean_text(ws.cell(row, 14).value),
            bidder_id=bidder_id,
            partner_id=partner_id,
            dealer_id=dealer_id,
            manual_review_required=session_number == "19",
        )
    for session_number, count in sequence_by_session.items():
        session = sessions[session_number]
        session["importedGameCount"] = count
        session["sourceSheetName"] = PRIMARY_SHEET
        session["qualityStatus"] = (
            "manual_review_required" if session_number == "19" else "ok"
        )
        session["missingScoreRows"] = 0
        session["cumulativeBlockStartColumn"] = 3
        session["deltaBlockStartColumn"] = 21
        session["preferredScoreBlockNumericRows"] = count
        session["headerRow"] = 3
        session["columnMapping"] = {
            "game_type_col": 14,
            "bidder_col": 8,
            "winner_col": 11,
            "dealer_col": 10,
            "partner_col": 9,
        }


def update_session_expected_counts(sessions: dict[str, dict[str, Any]]) -> None:
    accepted = {
        "3": 22,
        "4": 29,
        "8": 50,
    }
    for session_number, expected in accepted.items():
        if session_number in sessions:
            sessions[session_number]["acceptedExpectedGameCount"] = expected


def audit_summary(
    workbook: Any,
    games: list[dict[str, Any]],
    player_results: list[dict[str, Any]],
    sessions: list[dict[str, Any]],
    audit_result: dict[str, Any],
) -> dict[str, Any]:
    player_totals = Counter()
    for result in player_results:
        player_totals[result["playerId"]] += result["score"]
    field_counts = {
        "gameType": sum(1 for game in games if game["gameTypeNormalized"]),
        "dealer": sum(1 for game in games if game["dealerId"]),
        "bidder_or_winner": sum(1 for game in games if game["bidderId"] or game["winnerIds"]),
        "partner": sum(1 for game in games if game["partnerId"]),
        "score_sum_zero": sum(1 for game in games if game["checksum"] == 0),
    }
    issue_counts = audit_result.get("issueCounts", {})
    return {
        "version": "v3",
        "sheetCount": len(workbook.sheetnames),
        "importedSessions": len(sessions),
        "importedGames": len(games),
        "playerResultRows": len(player_results),
        "playerTotals": dict(sorted(player_totals.items())),
        "fieldCounts": field_counts,
        "issueCount": sum(issue_counts.values()),
        "issueCounts": issue_counts,
    }


def sha256_jsonable(value: Any) -> str:
    encoded = json.dumps(value, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def write_report(data: dict[str, Any], audit_result: dict[str, Any]) -> None:
    lines = [
        "# Whist import v3 - app-klart datasæt",
        "",
        f"Genereret: {data['generatedAt']}",
        "",
        "## Kort status",
        "",
        "| Måling | Antal |",
        "|---|---:|",
        f"| Sessions | {len(data['sessions'])} |",
        f"| Spil | {len(data['games'])} |",
        f"| PlayerResult-rækker | {len(data['playerResults'])} |",
        f"| Issues | {data['auditSummary']['issueCount']} |",
        "",
        "## Beslutninger",
        "",
        "- Session 1-4 og 6-7 importeres fra individuelle faner som begrænset kilde.",
        "- Session 5 bevares som tom kilde uden spil.",
        "- Session 19 inkluderes med `manual_review_required` quality flag.",
        "- Dato/sted kommer fra `00_Regnskab_01`; samlet-arkets dato er audit-only.",
        "",
        "## Issues fra audit",
        "",
        "| Issue | Antal |",
        "|---|---:|",
    ]
    for issue, count in sorted(audit_result.get("issueCounts", {}).items()):
        lines.append(f"| `{issue}` | {count} |")
    lines.append("")
    (ROOT / "IMPORT_V3_REPORT.md").write_text("\n".join(lines), encoding="utf-8")


def main(argv: list[str]) -> int:
    workbook_path = Path(argv[1]).expanduser() if len(argv) > 1 else audit_v3.DEFAULT_WORKBOOK
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 2
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    audit_result = audit_v3.audit(workbook_path)
    sessions_by_number = base_sessions(workbook, audit_result)
    games: list[dict[str, Any]] = []
    player_results: list[dict[str, Any]] = []
    import_individual_sheet_games(workbook, sessions_by_number, games, player_results)
    import_primary_sheet_games(workbook, sessions_by_number, games, player_results)
    update_session_expected_counts(sessions_by_number)

    sessions = sorted(
        sessions_by_number.values(),
        key=lambda session: audit_v3.sort_session_key(session["sessionNumber"]),
    )
    generated_at = datetime.now(timezone.utc).isoformat()
    data = {
        "version": "whist_historical_data_v3",
        "generatedAt": generated_at,
        "players": [
            {"id": player, "name": player, "displayOrder": index + 1, "isActive": True}
            for index, player in enumerate(PLAYERS)
        ],
        "sessions": sessions,
        "games": games,
        "playerResults": player_results,
        "auditSummary": audit_summary(
            workbook, games, player_results, sessions, audit_result
        ),
    }
    output_path = ROOT / "whist_historical_data_v3.json"
    output_path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    write_report(data, audit_result)
    manifest_path = ROOT / "import_manifest_v3.generated.json"
    manifest = {
        "bundleSemver": "3.0.0-draft",
        "generatedAtUtc": generated_at,
        "sourceWorkbook": {
            "fileName": workbook_path.name,
            "localPathAtGeneration": str(workbook_path),
            "sha256": audit_v3.sha256_file(workbook_path),
            "sheetCount": len(workbook.sheetnames),
        },
        "outputs": {
            "historicalDataJson": {
                "path": str(output_path.relative_to(ROOT.parent.parent.parent.parent)),
                "sha256": sha256_jsonable(data),
            }
        },
        "qualityPolicy": {
            "session19": "included_with_manual_review_required",
            "dateSource": AUDIT_SHEET,
        },
    }
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "sessions": len(sessions),
                "games": len(games),
                "playerResults": len(player_results),
                "output": str(output_path),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
