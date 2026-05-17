#!/usr/bin/env python3
"""Read-only audit for the v3 historical Whist workbook source.

This script intentionally does not generate app data. It verifies the workbook
revision, checks the v3 truth sheets, and emits audit artifacts that can be
reviewed before a real v3 importer is allowed to produce JSON.
"""

from __future__ import annotations

import csv
import hashlib
import json
import platform
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    import openpyxl
except ImportError:  # pragma: no cover - import above already exercises this.
    openpyxl = None


ROOT = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = (
    Path.home()
    / "Downloads"
    / "Whist – resultater – samlet (2024)_AKTIV_forenkling af data.xlsx"
)

PRIMARY_SHEET = "SAMLET_alle regnskab_16-5-2026"
AUDIT_SHEET = "00_Regnskab_01"
EXPECTED_WORKBOOK_SHA256 = (
    "a6c126c81f743c9ad5da54af3ac18e312459cc44d02e01d7b656d3ccce446bfa"
)
EXPECTED_PRIMARY_HEADER_SHA256 = (
    "81b3ab246b88c09f0d1b9ded92f1c3eec10fa612e3e76b3d6dd1f6f0bfcc4b77"
)

PLAYERS = ("Thomas", "Peter", "Janus", "Christian")
PLAYER_ALIASES = {
    "chrisitan": "Christian",
    "jan": "Janus",
    "janiz": "Janus",
    "janjus": "Janus",
    "janjusz": "Janus",
    "jansicz": "Janus",
    "jnaus": "Janus",
    "peer": "Peter",
    "thoms": "Thomas",
}
PLAYER_LOOKUP = {
    **{player.lower(): player for player in PLAYERS},
    **PLAYER_ALIASES,
}
SELF_PARTNER_MARKERS = {"selv makker", "selvmakker"}
INDIVIDUAL_SESSION_SHEETS = {
    "1": {
        "sheet": "01_21-06-2016",
        "status": "individual_sheet_limited_source",
        "deltaColumns": (11, 14),
        "cumulativeColumns": (2, 5),
    },
    "2": {
        "sheet": "02_2492016",
        "status": "individual_sheet_limited_source",
        "deltaColumns": (9, 12),
        "cumulativeColumns": (2, 5),
    },
    "3": {
        "sheet": "03_25-02-2017",
        "status": "individual_sheet_limited_source",
        "deltaColumns": (11, 14),
        "cumulativeColumns": (2, 5),
        "maxDataRow": 25,
        "acceptedExpectedGameCount": 22,
        "expectedGameCountNote": (
            "00_Regnskab_01 says 23, but sheet has 22 real games through row 25; "
            "later formula rows are ignored."
        ),
    },
    "4": {
        "sheet": "04_Måske Berlin",
        "status": "individual_sheet_limited_source",
        "deltaColumns": (11, 14),
        "cumulativeColumns": (2, 5),
        "acceptedExpectedGameCount": 29,
        "expectedGameCountNote": (
            "00_Regnskab_01 says 30, but sheet has 29 real games through row 32; "
            "final score matches the audit sheet."
        ),
    },
    "5": {
        "sheet": "05_22-09-2017_TOM",
        "status": "empty_source_sheet",
        "deltaColumns": None,
        "cumulativeColumns": None,
    },
    "6": {
        "sheet": "06_23-09-2017",
        "status": "individual_sheet_limited_source",
        "deltaColumns": None,
        "cumulativeColumns": (2, 5),
    },
    "7": {
        "sheet": "07_24-09-2017",
        "status": "individual_sheet_limited_source",
        "deltaColumns": None,
        "cumulativeColumns": (2, 5),
    },
}
ACCEPTED_PRIMARY_EXPECTED_GAME_COUNTS = {
    "8": {
        "acceptedExpectedGameCount": 50,
        "note": (
            "00_Regnskab_01 says 51, but primary sheet and 08_3-11-2018 contain "
            "50 real games; final score matches the audit sheet."
        ),
    },
}
KNOWN_CONTROL_TOTAL_ERRORS = {
    "27": {
        "player": "Christian",
        "controlValue": -72,
        "primaryValue": -76,
        "note": (
            "Known typo in 00_Regnskab_01. The primary sheet total is the "
            "accepted value; remove this exception once the workbook control "
            "sheet is corrected."
        ),
    },
}
SESSION19_SOURCE_SHEETS = [
    "19a_13-01-2023_Fredag",
    "19b_13-01-2023_Færge mod Tyskla",
    "19c_13-01-2023_Brewdog fredag",
]


@dataclass(frozen=True)
class Issue:
    severity: str
    issue: str
    sheet: str
    row: int | None
    session: str | None
    detail: str


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_session(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def is_session_key(value: str | None) -> bool:
    return value is not None and bool(value) and value[0].isdigit()


def normalize_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def normalize_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return None


def normalize_number(value: Any) -> int | float | None:
    if value in (None, ""):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    return None


def normalize_player(value: Any) -> str | None:
    if value in (None, ""):
        return None
    return PLAYER_LOOKUP.get(str(value).strip().lower())


def is_self_partner_marker(value: Any) -> bool:
    if value in (None, ""):
        return False
    return str(value).strip().lower() in SELF_PARTNER_MARKERS


def normalize_compound_bidder_partner(value: Any) -> tuple[str, str] | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    separator = "+" if "+" in text else "/" if "/" in text else None
    if separator is None:
        return None
    parts = [part.strip() for part in text.split(separator) if part.strip()]
    if len(parts) != 2:
        return None
    bidder = normalize_player(parts[0])
    partner = normalize_player(parts[1])
    if bidder is None or partner is None:
        return None
    return bidder, partner


def row_header_fingerprint(ws: Any, row_number: int) -> tuple[str, dict[str, str]]:
    values: list[str] = []
    mapped: dict[str, str] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row_number, col).value
        value = "" if raw is None else str(raw).strip()
        values.append(value)
        if value:
            mapped[get_column_letter(col)] = value
    digest = hashlib.sha256("\x1f".join(values).encode("utf-8")).hexdigest()
    return digest, mapped


def sort_session_key(session: str) -> tuple[float, str]:
    numeric = "".join(ch for ch in session if ch.isdigit() or ch == ".")
    try:
        return (float(numeric), session)
    except ValueError:
        return (9999, session)


def read_score_values(ws: Any, row: int, start_col: int, end_col: int) -> list[Any]:
    return [ws.cell(row, col).value for col in range(start_col, end_col + 1)]


def audit_individual_session_sheet(
    workbook: Any,
    session: str,
    config: dict[str, Any],
    issues: list[Issue],
) -> dict[str, Any]:
    sheet_name = config["sheet"]
    if sheet_name not in workbook.sheetnames:
        issues.append(
            Issue(
                "error",
                "missing_individual_session_sheet",
                sheet_name,
                None,
                session,
                "",
            )
        )
        return {
            "session": session,
            "sheet": sheet_name,
            "status": "missing_individual_session_sheet",
            "gameCount": 0,
            "missingScoreRows": 0,
            "scoreSumNotZeroRows": 0,
        }

    ws = workbook[sheet_name]
    delta_columns = config["deltaColumns"]
    cumulative_columns = config["cumulativeColumns"]
    max_data_row = config.get("maxDataRow", ws.max_row)
    game_count = 0
    missing_score_rows = 0
    score_sum_not_zero_rows = 0
    previous_cumulative: list[Any] | None = None

    for row in range(1, max_data_row + 1):
        marker = ws.cell(row, 1).value
        if not isinstance(marker, (int, float)) or row == 1:
            continue

        scores: list[Any]
        if delta_columns is not None:
            scores = read_score_values(ws, row, delta_columns[0], delta_columns[1])
            if not all(isinstance(value, (int, float)) for value in scores):
                scores = read_score_values(
                    ws, row, cumulative_columns[0], cumulative_columns[1]
                )
        elif cumulative_columns is not None:
            current = read_score_values(
                ws, row, cumulative_columns[0], cumulative_columns[1]
            )
            if previous_cumulative is None:
                scores = current
            else:
                scores = [
                    current[index] - previous_cumulative[index]
                    if isinstance(current[index], (int, float))
                    and isinstance(previous_cumulative[index], (int, float))
                    else None
                    for index in range(4)
                ]
            previous_cumulative = current
        else:
            continue

        if not all(isinstance(value, (int, float)) for value in scores):
            missing_score_rows += 1
            issues.append(
                Issue(
                    "warning",
                    "missing_individual_sheet_score_row",
                    sheet_name,
                    row,
                    session,
                    f"game marker {marker}",
                )
            )
            continue

        game_count += 1
        score_sum = sum(scores)
        if score_sum != 0:
            score_sum_not_zero_rows += 1
            issues.append(
                Issue(
                    "warning",
                    "score_sum_not_zero",
                    sheet_name,
                    row,
                    session,
                    f"game marker {marker}, score sum {score_sum}",
                )
            )

    issue_name = config["status"]
    issues.append(
        Issue(
            "info",
            issue_name,
            sheet_name,
            None,
            session,
            f"importable games {game_count}",
        )
    )
    return {
        "session": session,
        "sheet": sheet_name,
        "status": config["status"],
        "gameCount": game_count,
        "missingScoreRows": missing_score_rows,
        "scoreSumNotZeroRows": score_sum_not_zero_rows,
        "acceptedExpectedGameCount": config.get("acceptedExpectedGameCount"),
        "expectedGameCountNote": config.get("expectedGameCountNote"),
    }


def canonical_cumulative_scores(ws: Any, row: int) -> dict[str, Any] | None:
    scores: dict[str, Any] = {}
    for col in range(2, 6):
        player = normalize_player(ws.cell(2, col).value)
        value = ws.cell(row, col).value
        if player is None or not isinstance(value, (int, float)):
            return None
        scores[player] = value
    if set(scores) != set(PLAYERS):
        return None
    return scores


def build_session19_combined_rows(workbook: Any) -> list[dict[str, Any]]:
    source_rows: list[dict[str, Any]] = []
    for source_order, sheet_name in enumerate(SESSION19_SOURCE_SHEETS, start=1):
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        for row in range(3, ws.max_row + 1):
            marker = ws.cell(row, 1).value
            if not isinstance(marker, (int, float)):
                continue
            cumulative = canonical_cumulative_scores(ws, row)
            if cumulative is None:
                continue
            source_rows.append(
                {
                    "sourceOrder": source_order,
                    "sourceSheetName": sheet_name,
                    "sourceRow": row,
                    "sourceGameMarker": marker,
                    "meldingRaw": ws.cell(row, 7).value,
                    "bidderRaw": ws.cell(row, 8).value,
                    "partnerRaw": ws.cell(row, 9).value,
                    "dealerRaw": ws.cell(row, 10).value,
                    "cumulative": cumulative,
                }
            )

    source_rows.sort(
        key=lambda item: (
            float(item["sourceGameMarker"]),
            item["sourceOrder"],
            item["sourceRow"],
        )
    )

    marker_counts = Counter(float(row["sourceGameMarker"]) for row in source_rows)
    previous_cumulative: dict[str, Any] | None = None
    combined_rows: list[dict[str, Any]] = []
    for sequence, row in enumerate(source_rows, start=1):
        cumulative = row["cumulative"]
        if previous_cumulative is None:
            delta = cumulative
        else:
            delta = {
                player: cumulative[player] - previous_cumulative[player]
                for player in PLAYERS
            }
        previous_cumulative = cumulative
        score_sum = sum(delta[player] for player in PLAYERS)
        marker = float(row["sourceGameMarker"])
        combined_rows.append(
            {
                "sequenceInSession": sequence,
                "sourceGameMarker": normalize_session(row["sourceGameMarker"]),
                "duplicateSourceGameMarker": marker_counts[marker] > 1,
                "sourceSheetName": row["sourceSheetName"],
                "sourceRow": row["sourceRow"],
                "meldingRaw": row["meldingRaw"],
                "bidderRaw": row["bidderRaw"],
                "partnerRaw": row["partnerRaw"],
                "dealerRaw": row["dealerRaw"],
                "cumulativeThomas": cumulative["Thomas"],
                "cumulativePeter": cumulative["Peter"],
                "cumulativeJanus": cumulative["Janus"],
                "cumulativeChristian": cumulative["Christian"],
                "deltaThomas": delta["Thomas"],
                "deltaPeter": delta["Peter"],
                "deltaJanus": delta["Janus"],
                "deltaChristian": delta["Christian"],
                "deltaScoreSum": score_sum,
                "isZeroDeltaRow": all(delta[player] == 0 for player in PLAYERS),
            }
        )
    return combined_rows


def audit(workbook_path: Path) -> dict[str, Any]:
    issues: list[Issue] = []
    workbook_sha = sha256_file(workbook_path)
    if workbook_sha != EXPECTED_WORKBOOK_SHA256:
        issues.append(
            Issue(
                "error",
                "workbook_sha256_mismatch",
                "",
                None,
                None,
                f"expected {EXPECTED_WORKBOOK_SHA256}, got {workbook_sha}",
            )
        )

    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    sheet_names = workbook.sheetnames
    for required in (PRIMARY_SHEET, AUDIT_SHEET):
        if required not in sheet_names:
            issues.append(
                Issue("error", "missing_required_sheet", required, None, None, "")
            )

    if any(issue.issue == "missing_required_sheet" for issue in issues):
        return {
            "generatedAtUtc": datetime.now(timezone.utc).isoformat(),
            "workbookSha256": workbook_sha,
            "sheetCount": len(sheet_names),
            "issues": [issue.__dict__ for issue in issues],
        }

    primary = workbook[PRIMARY_SHEET]
    audit_sheet = workbook[AUDIT_SHEET]
    primary_header_sha, primary_headers = row_header_fingerprint(primary, 3)
    if primary_header_sha != EXPECTED_PRIMARY_HEADER_SHA256:
        issues.append(
            Issue(
                "error",
                "primary_header_sha256_mismatch",
                PRIMARY_SHEET,
                3,
                None,
                f"expected {EXPECTED_PRIMARY_HEADER_SHA256}, got {primary_header_sha}",
            )
        )

    control_sessions: dict[str, dict[str, Any]] = {}
    for row in range(3, audit_sheet.max_row + 1):
        session = normalize_session(audit_sheet.cell(row, 1).value)
        if not is_session_key(session):
            continue
        control_sessions[session] = {
            "row": row,
            "date": normalize_date(audit_sheet.cell(row, 3).value),
            "expectedGameCount": normalize_int(audit_sheet.cell(row, 9).value),
            "location": audit_sheet.cell(row, 12).value,
            "finalScores": {
                "Thomas": audit_sheet.cell(row, 4).value,
                "Peter": audit_sheet.cell(row, 5).value,
                "Janus": audit_sheet.cell(row, 6).value,
                "Christian": audit_sheet.cell(row, 7).value,
            },
            "scoreSumCheck": audit_sheet.cell(row, 8).value,
        }

    imported_counts: Counter[str] = Counter()
    imported_count_sources: dict[str, str] = {}
    score_zero_count = 0
    score_non_zero_count = 0
    dates_by_session: dict[str, set[str]] = defaultdict(set)
    unknown_names: Counter[str] = Counter()
    summary_only_sessions: set[str] = set()
    primary_final_scores: dict[str, dict[str, int | float | None]] = {}

    for row in range(4, primary.max_row + 1):
        session = normalize_session(primary.cell(row, 1).value)
        if session is None:
            continue
        primary_final_scores[session] = {
            player: normalize_number(primary.cell(row, col).value)
            for player, col in zip(PLAYERS, range(3, 7))
        }
        game_marker = primary.cell(row, 2).value
        if game_marker in (None, ""):
            summary_only_sessions.add(session)
            continue

        imported_counts[session] += 1
        row_date = normalize_date(primary.cell(row, 55).value)
        if row_date:
            dates_by_session[session].add(row_date)

        scores = [primary.cell(row, col).value for col in range(21, 25)]
        if all(isinstance(score, (int, float)) for score in scores):
            score_sum = sum(scores)
            if score_sum == 0:
                score_zero_count += 1
            else:
                score_non_zero_count += 1
                issues.append(
                    Issue(
                        "error",
                        "score_sum_not_zero",
                        PRIMARY_SHEET,
                        row,
                        session,
                        f"U:X sum is {score_sum}",
                    )
                )
        else:
            issues.append(
                Issue(
                    "error",
                    "missing_score_value",
                    PRIMARY_SHEET,
                    row,
                    session,
                    f"U:X values are {scores}",
                )
            )

        for label, col in (("Melder", 8), ("Makker", 9), ("giver", 10)):
            raw = primary.cell(row, col).value
            if raw in (None, ""):
                continue
            if label == "Makker" and is_self_partner_marker(raw):
                continue
            if label == "Melder" and normalize_compound_bidder_partner(raw):
                continue
            if label == "Melder" and is_self_partner_marker(raw):
                inferred = normalize_player(primary.cell(row, 10).value)
                if inferred is None:
                    unknown_names[f"{label}: {str(raw).strip()}"] += 1
                continue
            if normalize_player(raw) is None:
                value = str(raw).strip()
                unknown_names[f"{label}: {value}"] += 1

    primary_imported_counts = Counter(imported_counts)
    individual_sheet_audits: dict[str, dict[str, Any]] = {}
    for session in sorted(summary_only_sessions, key=sort_session_key):
        config = INDIVIDUAL_SESSION_SHEETS.get(session)
        if config is None:
            issues.append(
                Issue(
                    "info",
                    "summary_only_session",
                    PRIMARY_SHEET,
                    None,
                    session,
                    "No game marker in primary sheet; do not synthesize games.",
                )
            )
            continue
        sheet_audit = audit_individual_session_sheet(workbook, session, config, issues)
        individual_sheet_audits[session] = sheet_audit
        imported_counts[session] = sheet_audit["gameCount"]
        imported_count_sources[session] = sheet_audit["sheet"]

    for session in imported_counts:
        imported_count_sources.setdefault(session, PRIMARY_SHEET)

    individual_imported_games = sum(
        audit_row["gameCount"] for audit_row in individual_sheet_audits.values()
    )
    individual_missing_score_rows = sum(
        audit_row["missingScoreRows"] for audit_row in individual_sheet_audits.values()
    )
    individual_score_sum_not_zero_rows = sum(
        audit_row["scoreSumNotZeroRows"]
        for audit_row in individual_sheet_audits.values()
    )

    for session, audit_row in individual_sheet_audits.items():
        if audit_row["status"] == "empty_source_sheet":
            continue
        if audit_row["gameCount"] == 0:
            issues.append(
                Issue(
                    "warning",
                    "individual_source_no_games",
                    audit_row["sheet"],
                    None,
                    session,
                    "",
                )
            )

    for raw, count in sorted(unknown_names.items()):
        issues.append(
            Issue(
                "warning",
                "unknown_player_name",
                PRIMARY_SHEET,
                None,
                None,
                f"{raw} ({count} occurrences)",
            )
        )

    session_rows: list[dict[str, Any]] = []
    for session, control in sorted(control_sessions.items(), key=lambda item: sort_session_key(item[0])):
        imported = imported_counts.get(session, 0)
        expected = control["expectedGameCount"]
        dates = sorted(dates_by_session.get(session, set()))
        source = imported_count_sources.get(session, "")
        status = "ok"
        if session in individual_sheet_audits:
            status = individual_sheet_audits[session]["status"]
            if (
                status == "individual_sheet_limited_source"
                and individual_sheet_audits[session]["missingScoreRows"] > 0
            ):
                status = "individual_sheet_partial"
        elif session in summary_only_sessions and imported == 0:
            status = "summary_only"
        accepted_expected = individual_sheet_audits.get(session, {}).get(
            "acceptedExpectedGameCount"
        )
        if accepted_expected is None:
            accepted_expected = ACCEPTED_PRIMARY_EXPECTED_GAME_COUNTS.get(
                session, {}
            ).get("acceptedExpectedGameCount")
        if expected is not None and imported != expected and imported != accepted_expected:
            if status == "ok":
                status = "count_mismatch"
            elif "count_mismatch" not in status:
                status = f"{status}_count_mismatch"
            issues.append(
                Issue(
                    "warning",
                    "expected_vs_imported_count_mismatch",
                    source or PRIMARY_SHEET,
                    None,
                    session,
                    f"expected {expected}, imported {imported}",
                )
            )
        if control["date"] and dates and control["date"] not in dates:
            issues.append(
                Issue(
                    "warning",
                    "date_mismatch",
                    PRIMARY_SHEET,
                    None,
                    session,
                    f"control date {control['date']}, primary dates {', '.join(dates)}",
                )
            )
        primary_final = primary_final_scores.get(session)
        control_final = {
            player: normalize_number(value)
            for player, value in control["finalScores"].items()
        }
        total_status = "not_checked"
        if primary_final and all(control_final[player] is not None for player in PLAYERS):
            differences = {
                player: (primary_final[player] or 0) - (control_final[player] or 0)
                for player in PLAYERS
            }
            non_zero_differences = {
                player: value for player, value in differences.items() if value != 0
            }
            total_status = "ok"
            if non_zero_differences:
                known_error = KNOWN_CONTROL_TOTAL_ERRORS.get(session)
                known_error_matches = (
                    known_error is not None
                    and set(non_zero_differences) == {known_error["player"]}
                    and control_final[known_error["player"]] == known_error["controlValue"]
                    and primary_final[known_error["player"]] == known_error["primaryValue"]
                )
                if known_error_matches:
                    total_status = "known_control_total_error"
                    issues.append(
                        Issue(
                            "info",
                            "known_control_total_error",
                            AUDIT_SHEET,
                            control["row"],
                            session,
                            f"{known_error['player']} control {known_error['controlValue']} "
                            f"vs primary {known_error['primaryValue']}. {known_error['note']}",
                        )
                    )
                else:
                    total_status = "control_total_mismatch"
                    issues.append(
                        Issue(
                            "warning",
                            "control_total_mismatch",
                            AUDIT_SHEET,
                            control["row"],
                            session,
                            "control vs primary final score diff "
                            + ", ".join(
                                f"{player}: {value}"
                                for player, value in non_zero_differences.items()
                            ),
                        )
                    )
        if session == "19":
            status = "manual_review_required"
        session_rows.append(
            {
                "session": session,
                "controlDate": control["date"],
                "primaryDates": ", ".join(dates),
                "expectedGameCount": expected,
                "acceptedExpectedGameCount": accepted_expected,
                "importedGameCount": imported,
                "importedCountSource": source,
                "location": control["location"],
                "status": status,
                "totalReconciliationStatus": total_status,
            }
        )

    issue_counts = Counter(issue.issue for issue in issues)
    session19_combined_rows = build_session19_combined_rows(workbook)
    if session19_combined_rows:
        duplicate_markers = sorted(
            {
                row["sourceGameMarker"]
                for row in session19_combined_rows
                if row["duplicateSourceGameMarker"]
            },
            key=lambda marker: float(marker),
        )
        zero_delta_sequences = [
            str(row["sequenceInSession"])
            for row in session19_combined_rows
            if row["isZeroDeltaRow"]
        ]
        issues.append(
            Issue(
                "warning",
                "session19_manual_review_required",
                "session19_combined_order.csv",
                None,
                "19",
                "duplicate markers "
                + ", ".join(duplicate_markers)
                + "; zero-delta sequences "
                + ", ".join(zero_delta_sequences),
            )
        )
        issue_counts = Counter(issue.issue for issue in issues)
    return {
        "generatedAtUtc": datetime.now(timezone.utc).isoformat(),
        "workbook": {
            "path": str(workbook_path),
            "sha256": workbook_sha,
            "sheetCount": len(sheet_names),
        },
        "sheets": {
            PRIMARY_SHEET: {
                "rows": primary.max_row,
                "columns": primary.max_column,
                "headerRow": 3,
                "headerSha256": primary_header_sha,
                "headers": primary_headers,
                "summaryOnlyRows": len(summary_only_sessions),
                "gameRows": sum(primary_imported_counts.values()),
                "sessionsWithGameRows": len(primary_imported_counts),
                "scoreSumZeroRows": score_zero_count,
                "scoreSumNonZeroRows": score_non_zero_count,
            },
            AUDIT_SHEET: {
                "rows": audit_sheet.max_row,
                "columns": audit_sheet.max_column,
                "controlSessions": len(control_sessions),
            },
        },
        "individualSessionSheets": {
            "importedGames": individual_imported_games,
            "missingScoreRows": individual_missing_score_rows,
            "scoreSumNotZeroRows": individual_score_sum_not_zero_rows,
            "sessions": individual_sheet_audits,
        },
        "combinedCoverage": {
            "importedGames": sum(imported_counts.values()),
            "sessionsWithImportedGames": sum(1 for count in imported_counts.values() if count > 0),
        },
        "session19Combined": {
            "rows": len(session19_combined_rows),
            "uniqueSourceGameMarkers": len(
                {row["sourceGameMarker"] for row in session19_combined_rows}
            ),
            "duplicateSourceGameMarkers": sorted(
                {
                    row["sourceGameMarker"]
                    for row in session19_combined_rows
                    if row["duplicateSourceGameMarker"]
                },
                key=lambda marker: float(marker),
            ),
            "zeroDeltaRows": sum(
                1 for row in session19_combined_rows if row["isZeroDeltaRow"]
            ),
        },
        "sessions": session_rows,
        "session19CombinedRows": session19_combined_rows,
        "issueCounts": dict(sorted(issue_counts.items())),
        "issues": [issue.__dict__ for issue in issues],
        "tooling": {
            "pythonVersion": platform.python_version(),
            "openpyxlVersion": getattr(openpyxl, "__version__", None),
        },
    }


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_report(path: Path, result: dict[str, Any]) -> None:
    primary = result["sheets"][PRIMARY_SHEET]
    audit_sheet = result["sheets"][AUDIT_SHEET]
    individual = result["individualSessionSheets"]
    combined = result["combinedCoverage"]
    session19 = result["session19Combined"]
    lines = [
        "# Whist import v3 - read-only audit",
        "",
        f"Genereret: {result['generatedAtUtc']}",
        "",
        "## Workbook",
        "",
        f"- SHA256: `{result['workbook']['sha256']}`",
        f"- Faner: {result['workbook']['sheetCount']}",
        f"- Python: `{result['tooling']['pythonVersion']}`",
        f"- openpyxl: `{result['tooling']['openpyxlVersion']}`",
        "",
        "## Arkstatus",
        "",
        "| Ark | Rækker | Kolonner | Rolle |",
        "|---|---:|---:|---|",
        f"| `{PRIMARY_SHEET}` | {primary['rows']} | {primary['columns']} | Primær spil-for-spil-kilde |",
        f"| `{AUDIT_SHEET}` | {audit_sheet['rows']} | {audit_sheet['columns']} | Audit/kontrol |",
        "",
        "## Primærkilde",
        "",
        f"- Header SHA256 række 3: `{primary['headerSha256']}`",
        f"- Summary/status-rækker for session 1-7: {primary['summaryOnlyRows']}",
        f"- Spilrækker: {primary['gameRows']}",
        f"- Sessions med spilrækker: {primary['sessionsWithGameRows']}",
        f"- Score-sum = 0: {primary['scoreSumZeroRows']}",
        f"- Score-sum != 0: {primary['scoreSumNonZeroRows']}",
        "",
        "## Samlet v3-dækning",
        "",
        f"- Importerbare spil på tværs af kilder: {combined['importedGames']}",
        f"- Sessions med importerbare spil: {combined['sessionsWithImportedGames']}",
        "",
        "## Session 19 samlet rækkefølge",
        "",
        f"- Rækker: {session19['rows']}",
        f"- Unikke spilnumre fra kolonne A: {session19['uniqueSourceGameMarkers']}",
        f"- Dublet-spilnumre: {', '.join(session19['duplicateSourceGameMarkers'])}",
        f"- Nul-delta-rækker: {session19['zeroDeltaRows']}",
        "- CSV: `session19_combined_order.csv`",
        "",
        "## Individuelle faner 1-7",
        "",
        f"- Importerbare spil: {individual['importedGames']}",
        f"- Manglende score-rækker: {individual['missingScoreRows']}",
        f"- Score-sum != 0: {individual['scoreSumNotZeroRows']}",
        "",
        "| Session | Fane | Importerbare spil | Accepteret forventet | Manglende score-rækker | Score-sum != 0 | Status |",
        "|---|---|---:|---:|---:|---:|---|",
    ]
    for session, row in sorted(
        individual["sessions"].items(), key=lambda item: sort_session_key(item[0])
    ):
        lines.append(
            "| {session} | `{sheet}` | {gameCount} | {acceptedExpectedGameCount} | {missingScoreRows} | {scoreSumNotZeroRows} | {status} |".format(
                **{
                    **row,
                    "acceptedExpectedGameCount": row.get("acceptedExpectedGameCount")
                    or "",
                }
            )
        )
    lines.extend(
        [
        "",
        "## Issues",
        "",
        "| Issue | Antal |",
        "|---|---:|",
        ]
    )
    for issue, count in result["issueCounts"].items():
        lines.append(f"| `{issue}` | {count} |")
    lines.extend(
        [
            "",
            "## Sessions",
            "",
            "| Session | Dato kontrol | Dato primær | Forventet | Accepteret forventet | Importeret | Kilde | Status | Totalafstemning |",
            "|---|---|---|---:|---:|---:|---|---|---|",
        ]
    )
    for row in result["sessions"]:
        expected = "" if row["expectedGameCount"] is None else row["expectedGameCount"]
        lines.append(
            "| {session} | {controlDate} | {primaryDates} | {expected} | {acceptedExpected} | {importedGameCount} | {importedCountSource} | {status} | {totalReconciliationStatus} |".format(
                expected=expected,
                acceptedExpected=row.get("acceptedExpectedGameCount") or "",
                **row,
            )
        )
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def main(argv: list[str]) -> int:
    workbook_path = Path(argv[1]).expanduser() if len(argv) > 1 else DEFAULT_WORKBOOK
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 2

    result = audit(workbook_path)
    (ROOT / "import_manifest_v3.audit.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    write_csv(
        ROOT / "session_validation_v3.csv",
        result.get("sessions", []),
        [
            "session",
            "controlDate",
            "primaryDates",
            "expectedGameCount",
            "acceptedExpectedGameCount",
            "importedGameCount",
            "importedCountSource",
            "location",
            "status",
            "totalReconciliationStatus",
        ],
    )
    write_csv(
        ROOT / "session19_combined_order.csv",
        result.get("session19CombinedRows", []),
        [
            "sequenceInSession",
            "sourceGameMarker",
            "duplicateSourceGameMarker",
            "sourceSheetName",
            "sourceRow",
            "meldingRaw",
            "bidderRaw",
            "partnerRaw",
            "dealerRaw",
            "cumulativeThomas",
            "cumulativePeter",
            "cumulativeJanus",
            "cumulativeChristian",
            "deltaThomas",
            "deltaPeter",
            "deltaJanus",
            "deltaChristian",
            "deltaScoreSum",
            "isZeroDeltaRow",
        ],
    )
    write_csv(
        ROOT / "issues_v3.csv",
        result.get("issues", []),
        ["severity", "issue", "sheet", "row", "session", "detail"],
    )
    if PRIMARY_SHEET in result.get("sheets", {}):
        write_report(ROOT / "IMPORT_V3_AUDIT.md", result)
    print(json.dumps({"issues": result.get("issueCounts", {}), "outputDir": str(ROOT)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
