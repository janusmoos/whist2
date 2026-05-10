import json
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
WORKBOOK_PATH = ROOT / "docs/statistik/examples/Whist – resultater – samlet (2024)_AKTIV_forenkling af data.xlsx"
JSON_PATH = ROOT / "Whist20/Resources/HistoricalData/whist_historical_data_v2.json"
REPORT_PATH = ROOT / "docs/statistik/session26_reimport_analysis.md"
PLAYERS = ["Thomas", "Peter", "Janus", "Christian"]


def score_text(values):
    return ", ".join(f"{player} {value:+d}" for player, value in zip(PLAYERS, values))


def score_text_from_dict(scores):
    return ", ".join(f"{player} {scores[player]:+d}" for player in PLAYERS)


def numeric_scores(sheet, row, start_column):
    values = []
    for offset in range(4):
        value = sheet.cell(row=row, column=start_column + offset).value
        if not isinstance(value, (int, float)) or isinstance(value, bool):
            return None
        values.append(int(round(value)))
    return values


def remove_flag(flags, flag):
    return [item for item in flags if item != flag]


def recompute_audit_summary(data):
    games = data["games"]
    results = data["playerResults"]
    player_totals = {player: 0 for player in PLAYERS}
    for result in results:
        if result["playerId"] in player_totals:
            player_totals[result["playerId"]] += result["score"]

    issue_counts = Counter()
    for game in games:
        for flag in game.get("qualityFlags", []):
            if flag in ("score_sum_not_zero", "source_explicit_score_sum_not_zero"):
                issue_counts[flag] += 1

    old_issue_counts = data.get("auditSummary", {}).get("issueCounts", {})
    for preserved in ("game_marker_without_scores", "expected_vs_imported_count_mismatch"):
        if preserved in old_issue_counts:
            issue_counts[preserved] = old_issue_counts[preserved]

    field_counts = {
        "gameType": sum(1 for game in games if game.get("gameTypeRaw")),
        "dealer": sum(1 for game in games if game.get("dealerId")),
        "bidder_or_winner": sum(1 for game in games if game.get("bidderId") or game.get("winnerId")),
        "partner": sum(1 for game in games if game.get("partnerId")),
        "score_sum_zero": sum(1 for game in games if game.get("checksum") == 0),
    }

    audit = data["auditSummary"]
    audit["playerTotals"] = player_totals
    audit["fieldCounts"] = field_counts
    audit["issueCounts"] = dict(sorted(issue_counts.items()))
    audit["issueCount"] = sum(issue_counts.values())


def is_team_score_mismatch(game, scores):
    partner = (game.get("partnerId") or "").strip()
    game_type = (game.get("gameTypeNormalized") or game.get("gameTypeRaw") or "").strip().lower()
    if not partner or partner == "Selvmakker":
        return False
    if any(marker in game_type for marker in ("sol", "bordlægger", "storslem")):
        return False
    if partner not in PLAYERS:
        return False

    anchors = [player for player in game.get("bidderIds", []) if player in PLAYERS]
    if not anchors and game.get("bidderId") in PLAYERS:
        anchors = [game["bidderId"]]
    if not anchors and game.get("winnerId") in PLAYERS:
        anchors = [game["winnerId"]]
    if len(anchors) != 1 or anchors[0] == partner:
        return False

    team = [anchors[0], partner]
    opponents = [player for player in PLAYERS if player not in team]
    return scores[team[0]] != scores[team[1]] or scores[opponents[0]] != scores[opponents[1]]


def main():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True, read_only=False)
    sheet = workbook["26_30-11-2024"]
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))

    results_by_game = {}
    for result in data["playerResults"]:
        results_by_game.setdefault(result["gameId"], {})[result["playerId"]] = result

    corrections = []
    skipped = []

    for game in data["games"]:
        if game["sourceSheetName"] != "26_30-11-2024":
            continue

        row = game["sourceRow"]
        explicit_scores = numeric_scores(sheet, row, 19)  # S:V, labelled T j / P j / J j / C j.
        current_scores = [results_by_game[game["id"]][player]["score"] for player in PLAYERS]

        if explicit_scores is None:
            skipped.append((game, current_scores, None, "mangler numerisk scoreblok S:V"))
            continue

        if sum(explicit_scores) != 0:
            flag = "source_explicit_score_sum_not_zero"
            if flag not in game["qualityFlags"]:
                game["qualityFlags"].append(flag)
            skipped.append((game, current_scores, explicit_scores, "S:V summerer ikke til nul"))
            continue

        if explicit_scores == current_scores:
            game["scoreSource"] = "corrected_explicit_score_block_S_V"
            continue

        for player, score in zip(PLAYERS, explicit_scores):
            results_by_game[game["id"]][player]["score"] = score

        game["checksum"] = sum(explicit_scores)
        game["scoreSource"] = "corrected_explicit_score_block_S_V"
        game["qualityFlags"] = remove_flag(game["qualityFlags"], "score_sum_not_zero")
        corrections.append((game, current_scores, explicit_scores))

    data["generatedAt"] = datetime.now().isoformat(timespec="seconds")
    data["version"] = "whist_historical_data_v2_corrected"
    recompute_audit_summary(data)

    corrected_scores_by_game = {}
    for game in data["games"]:
        corrected_scores_by_game[game["id"]] = {
            player: results_by_game[game["id"]][player]["score"]
            for player in PLAYERS
            if game["id"] in results_by_game and player in results_by_game[game["id"]]
        }

    remaining_team_mismatches = [
        game for game in data["games"]
        if set(corrected_scores_by_game.get(game["id"], {})) == set(PLAYERS)
        and is_team_score_mismatch(game, corrected_scores_by_game[game["id"]])
    ]

    JSON_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    report_lines = [
        "# Spilledag 26 – analyse og korrigeret reimport\n",
        "\n",
        f"Genereret: {data['generatedAt']}\n",
        "\n",
        "## Konklusion\n",
        "\n",
        "Spilledag 26 blev tidligere importeret fra forskellen mellem kolonne B:E-rækkerne. "
        "Regnearket indeholder imidlertid en eksplicit per-spil-scoreblok i kolonne S:V "
        "(`T j`, `P j`, `J j`, `C j`). Den blok er brugt som korrigeret kilde, når den summerer til nul.\n",
        "\n",
        "Spil 3 (`Sang`) var derfor forkert importeret som `Thomas -92, Peter +84, Janus +84, Christian -76`. "
        "Den korrigerede reimport bruger `Thomas -80, Peter +80, Janus +80, Christian -80`.\n",
        "\n",
        "## Korrigerede spil\n",
        "\n",
        "| Spil | Type | Før | Efter |\n",
        "|---:|---|---|---|\n",
    ]

    for game, before, after in corrections:
        report_lines.append(
            f"| {game['gameNumberInSession']} | {game.get('gameTypeRaw') or '-'} | "
            f"{score_text(before)} | {score_text(after)} |\n"
        )

    report_lines.extend([
        "\n",
        "## Ikke automatisk korrigeret\n",
        "\n",
        "| Spil | Type | Importeret | Regneark S:V | Årsag |\n",
        "|---:|---|---|---|---|\n",
    ])

    for game, current, explicit, reason in skipped:
        explicit_text = "-" if explicit is None else score_text(explicit)
        report_lines.append(
            f"| {game['gameNumberInSession']} | {game.get('gameTypeRaw') or '-'} | "
            f"{score_text(current)} | {explicit_text} | {reason} |\n"
        )

    report_lines.extend([
        "\n",
        "## Andre holdscore-afvigelser i historikken\n",
        "\n",
        "De øvrige makkerspil med ulige holdscore er ikke automatisk ændret. "
        "Ved opslag i regnearket findes samme afvigelse i kildeblokken, så de er markeret som datakvalitet/regelfortolkning "
        "snarere end en sikker importfejl.\n",
        "\n",
        "| Spilledag | Spil | Type | Resultat efter reimport | Vurdering |\n",
        "|---:|---:|---|---|---|\n",
    ])

    for game in remaining_team_mismatches:
        report_lines.append(
            f"| {game['sessionNumber']} | {game['gameNumberInSession']} | {game.get('gameTypeRaw') or '-'} | "
            f"{score_text_from_dict(corrected_scores_by_game[game['id']])} | Kilde-/regelafklaring; ikke sikker importfejl |\n"
        )

    REPORT_PATH.write_text("".join(report_lines), encoding="utf-8")
    print(
        f"Corrected {len(corrections)} games; skipped {len(skipped)} games; "
        f"remaining team mismatches {len(remaining_team_mismatches)}."
    )


if __name__ == "__main__":
    main()
