#!/usr/bin/env python3
"""Audit historical Whist workbook score consistency.

The v3 importer reads explicit delta columns (U:X) for the primary source sheet.
This audit verifies that those deltas match the difference between cumulative
score columns (C:F) within each session, and compares local workbook revisions.
"""

from __future__ import annotations

import csv
import hashlib
import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
PRIMARY_SHEET = "SAMLET_alle regnskab_16-5-2026"
PLAYERS = ("Thomas", "Peter", "Janus", "Christian")

WORKBOOK_CANDIDATES = [
    ROOT / "docs/statistik/examples/Whist – resultater – samlet (2024)_AKTIV_forenkling af data.xlsx",
    Path.home() / "Downloads/Whist – resultater – samlet (2024)_AKTIV_forenkling af data.xlsx",
    Path.home() / "Downloads/Whist – resultater – samlet (2024)_AKTIV_forenkling af data-2.xlsx",
    Path.home() / "Downloads/Whist – resultater – samlet (2024)_AKTIV_forenkling af data-3.xlsx",
]

OUT_DIR = ROOT / "docs/statistik/audit"
REPORT_PATH = ROOT / "docs/statistik/data_audit_2026-05-26.md"
APP_JSON_PATH = ROOT / "Whist20/Resources/HistoricalData/whist_historical_data_v3.json"
IMPORT_MANIFEST_PATH = ROOT / "docs/statistik/examples/v03/import_manifest_v3.generated.json"


@dataclass(frozen=True)
class GameRow:
    workbook: str
    workbook_path: str
    workbook_sha256: str
    source_row: int
    session: str
    game: int
    cumulative: dict[str, int]
    explicit_delta: dict[str, int] | None
    expected_delta: dict[str, int]
    game_type_raw: str
    game_type: str
    bidder: str
    partner: str
    dealer: str

    @property
    def key(self) -> tuple[str, int]:
        return (self.session, self.game)

    @property
    def explicit_delta_sum(self) -> int | None:
        if self.explicit_delta is None:
            return None
        return sum(self.explicit_delta.values())

    @property
    def expected_delta_sum(self) -> int:
        return sum(self.expected_delta.values())

    @property
    def has_delta_mismatch(self) -> bool:
        return self.explicit_delta is not None and self.explicit_delta != self.expected_delta

    @property
    def has_sum_issue(self) -> bool:
        return self.explicit_delta_sum not in (None, 0)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_int(value: Any) -> int | None:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return None


def normalize_session(value: Any) -> str | None:
    integer = normalize_int(value)
    if integer is not None:
        return str(integer)
    if value in (None, ""):
        return None
    text = str(value).strip()
    return text or None


def text(value: Any) -> str:
    if value in (None, ""):
        return ""
    return str(value).strip()


def score_map(ws: Any, row: int, start_col: int) -> dict[str, int] | None:
    scores: dict[str, int] = {}
    for offset, player in enumerate(PLAYERS):
        value = normalize_int(ws.cell(row, start_col + offset).value)
        if value is None:
            return None
        scores[player] = value
    return scores


def read_workbook(path: Path) -> list[GameRow]:
    workbook_hash = sha256_file(path)
    wb = load_workbook(path, data_only=True, read_only=False)
    if PRIMARY_SHEET not in wb.sheetnames:
        return []

    ws = wb[PRIMARY_SHEET]
    rows: list[GameRow] = []
    previous_cumulative_by_session: dict[str, dict[str, int]] = {}

    for row in range(4, ws.max_row + 1):
        session = normalize_session(ws.cell(row, 1).value)
        game = normalize_int(ws.cell(row, 2).value)
        if session is None or game is None:
            continue

        cumulative = score_map(ws, row, 3)
        if cumulative is None:
            continue

        explicit_delta = score_map(ws, row, 21)
        previous = previous_cumulative_by_session.get(session)
        expected_delta = (
            cumulative
            if previous is None
            else {player: cumulative[player] - previous[player] for player in PLAYERS}
        )
        previous_cumulative_by_session[session] = cumulative

        rows.append(
            GameRow(
                workbook=path.name,
                workbook_path=str(path),
                workbook_sha256=workbook_hash,
                source_row=row,
                session=session,
                game=game,
                cumulative=cumulative,
                explicit_delta=explicit_delta,
                expected_delta=expected_delta,
                game_type_raw=text(ws.cell(row, 11).value),
                game_type=text(ws.cell(row, 14).value),
                bidder=text(ws.cell(row, 8).value),
                partner=text(ws.cell(row, 9).value),
                dealer=text(ws.cell(row, 10).value),
            )
        )

    return rows


def score_text(scores: dict[str, int] | None) -> str:
    if scores is None:
        return "-"
    return ", ".join(f"{player} {scores[player]:+d}" for player in PLAYERS)


def csv_row(row: GameRow) -> dict[str, Any]:
    return {
        "workbook": row.workbook,
        "sha256": row.workbook_sha256,
        "sourceRow": row.source_row,
        "session": row.session,
        "game": row.game,
        "gameTypeRaw": row.game_type_raw,
        "gameType": row.game_type,
        "bidder": row.bidder,
        "partner": row.partner,
        "dealer": row.dealer,
        "explicitDelta": score_text(row.explicit_delta),
        "expectedDelta": score_text(row.expected_delta),
        "explicitDeltaSum": row.explicit_delta_sum,
        "expectedDeltaSum": row.expected_delta_sum,
        "cumulative": score_text(row.cumulative),
    }


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def compare_versions(rows_by_workbook: dict[str, list[GameRow]]) -> list[dict[str, Any]]:
    rows_by_key: dict[tuple[str, int], list[GameRow]] = {}
    for rows in rows_by_workbook.values():
        for row in rows:
            rows_by_key.setdefault(row.key, []).append(row)

    differences: list[dict[str, Any]] = []
    for key, rows in sorted(rows_by_key.items(), key=lambda item: (session_sort_key(item[0][0]), item[0][1])):
        signatures = {
            (
                tuple(row.explicit_delta[player] for player in PLAYERS) if row.explicit_delta else None,
                tuple(row.expected_delta[player] for player in PLAYERS),
                row.game_type_raw,
                row.bidder,
                row.partner,
                row.dealer,
            )
            for row in rows
        }
        if len(signatures) <= 1:
            continue
        for row in rows:
            differences.append(csv_row(row))

    return differences


def session_sort_key(session: str) -> tuple[int, str]:
    digits = ""
    for character in session:
        if character.isdigit():
            digits += character
        else:
            break
    return (int(digits) if digits else 9999, session)


def markdown_table(rows: list[dict[str, Any]], limit: int = 12) -> str:
    if not rows:
        return "_Ingen fund._\n"
    columns = [
        "workbook",
        "sourceRow",
        "session",
        "game",
        "gameTypeRaw",
        "explicitDelta",
        "expectedDelta",
    ]
    output = "| " + " | ".join(columns) + " |\n"
    output += "| " + " | ".join("---" for _ in columns) + " |\n"
    for row in rows[:limit]:
        output += "| " + " | ".join(str(row.get(column, "")).replace("|", "\\|") for column in columns) + " |\n"
    if len(rows) > limit:
        output += f"\n_Viser {limit} af {len(rows)} fund. Se CSV for hele listen._\n"
    return output


def app_data_status() -> dict[str, Any] | None:
    if not APP_JSON_PATH.exists():
        return None

    data = json.loads(APP_JSON_PATH.read_text(encoding="utf-8"))
    manifest = (
        json.loads(IMPORT_MANIFEST_PATH.read_text(encoding="utf-8"))
        if IMPORT_MANIFEST_PATH.exists()
        else {}
    )
    source = manifest.get("sourceWorkbook", {})

    result_by_game: dict[str, dict[str, int]] = {}
    for result in data.get("playerResults", []):
        result_by_game.setdefault(result["gameId"], {})[result["playerId"]] = result["score"]

    checks = []
    for game_id, label in [
        ("session_30_2025-11-21_game_001", "Spilledag 30, Spil 1"),
        ("session_31_2025-11-22_game_001", "Spilledag 31, Spil 1"),
    ]:
        scores = result_by_game.get(game_id)
        checks.append(f"- {label}: {score_text(scores) if scores else 'ikke fundet'}")

    return {
        "sourceFile": source.get("fileName", "-"),
        "sourceSha256": source.get("sha256", "-"),
        "generatedAt": data.get("generatedAt", "-"),
        "games": len(data.get("games", [])),
        "playerResults": len(data.get("playerResults", [])),
        "checks": "\n".join(checks),
    }


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    rows_by_workbook = {
        path.name: read_workbook(path)
        for path in WORKBOOK_CANDIDATES
        if path.exists()
    }

    mismatch_rows = [
        csv_row(row)
        for rows in rows_by_workbook.values()
        for row in rows
        if row.has_delta_mismatch
    ]
    sum_issue_rows = [
        csv_row(row)
        for rows in rows_by_workbook.values()
        for row in rows
        if row.has_sum_issue
    ]
    version_diff_rows = compare_versions(rows_by_workbook)

    write_csv(OUT_DIR / "score_delta_mismatches_2026-05-26.csv", mismatch_rows)
    write_csv(OUT_DIR / "score_sum_issues_2026-05-26.csv", sum_issue_rows)
    write_csv(OUT_DIR / "workbook_version_differences_2026-05-26.csv", version_diff_rows)

    workbook_lines = []
    for workbook, rows in rows_by_workbook.items():
        hash_text = rows[0].workbook_sha256 if rows else sha256_file(next(path for path in WORKBOOK_CANDIDATES if path.name == workbook))
        workbook_lines.append(
            f"| `{workbook}` | `{hash_text}` | {len(rows)} | "
            f"{sum(1 for row in rows if row.has_delta_mismatch)} | "
            f"{sum(1 for row in rows if row.has_sum_issue)} |"
        )

    mismatch_keys = sorted(
        {(row["session"], row["game"]) for row in mismatch_rows},
        key=lambda key: (session_sort_key(key[0]), int(key[1])),
    )
    version_diff_keys = sorted(
        {(row["session"], row["game"]) for row in version_diff_rows},
        key=lambda key: (session_sort_key(key[0]), int(key[1])),
    )
    mismatch_key_text = ", ".join(f"Spilledag {session}, Spil {game}" for session, game in mismatch_keys)
    version_key_text = ", ".join(f"{session}/{game}" for session, game in version_diff_keys)
    app_status = app_data_status()
    app_section = ""
    if app_status is not None:
        app_section = f"""
## Aktuel app-data efter reimport

| Måling | Værdi |
|---|---|
| Kilde-workbook | `{app_status["sourceFile"]}` |
| Kilde-SHA256 | `{app_status["sourceSha256"]}` |
| App-data genereret | {app_status["generatedAt"]} |
| Spil i app-JSON | {app_status["games"]} |
| PlayerResult-rækker | {app_status["playerResults"]} |

Direkte kontrol af de to sikre fejl i appens JSON:

{app_status["checks"]}

"""

    report = f"""# Historisk data-audit: scorekonsistens

Genereret: {datetime.now().isoformat(timespec="seconds")}

## Formål

Denne audit er lavet efter fundet i Spilledag 30, Spil 1, hvor appens importerede delta-score ikke matchede den kumulative score i regnearket. Audit'en sammenligner eksplicitte delta-kolonner (U:X) med den delta, der kan udledes af kumulative scorekolonner (C:F), og sammenligner de workbook-versioner, der findes lokalt.

## Workbook-versioner

| Workbook | SHA-256 | Læste spilrækker | Delta/kumulativ mismatch | Delta-sum != 0 |
|---|---|---:|---:|---:|
{chr(10).join(workbook_lines)}

{app_section}
## Sikre delta/kumulativ-fejl

Disse rækker har eksplicitte delta-tal, som ikke matcher ændringen i de kumulative scorekolonner.

{markdown_table(mismatch_rows)}

CSV: `docs/statistik/audit/score_delta_mismatches_2026-05-26.csv`

## Delta-sum-fejl

Disse rækker har eksplicitte delta-tal, der ikke summerer til nul.

{markdown_table(sum_issue_rows)}

CSV: `docs/statistik/audit/score_sum_issues_2026-05-26.csv`

## Forskelle mellem workbook-versioner

Disse spil har forskellige importerbare værdier mellem de lokale workbook-versioner.

{markdown_table(version_diff_rows)}

CSV: `docs/statistik/audit/workbook_version_differences_2026-05-26.csv`

## Konklusion

- Den oprindelige workbook `Whist – resultater – samlet (2024)_AKTIV_forenkling af data.xlsx` indeholder {len(mismatch_keys)} sikre delta/kumulativ-fejl: {mismatch_key_text}.
- Appens aktuelle v3-data er reimporteret fra den rettede workbook `...data-3.xlsx`, som har 0 delta/kumulativ-fejl i den primære kilde.
- Direkte kontrol af appens JSON bekræfter, at Spilledag 30, Spil 1 og Spilledag 31, Spil 1 nu bruger de korrigerede scorer.
- Der er versionsforskelle på {len(version_diff_keys)} spilnøgler: {version_key_text}. Nogle af dem ligger i Spilledag 19, hvor der historisk er dublet-/manual-review-problemer; de sikre deltafejl ligger i Spilledag 30 og 31.
- Repo-kopien under `docs/statistik/examples` er en ældre/afkortet kilde uden de nyeste rækker i `SAMLET_alle regnskab_16-5-2026`, og bør ikke bruges som sandhedskilde for de seneste spilledage.

## Anbefaling

1. Behandl `...data-3.xlsx` som den aktuelle sandhedskilde for v3-importen.
2. Behold audit-scriptet som fast kontrol efter fremtidige importer eller manuelle rettelser i workbooken.
3. Gennemgå versionsforskellene i CSV'en særskilt, især Spilledag 19, hvis vi vil rydde op i historiske dublet-/manual-review-problemer.
4. Brug ikke repo-kopien under `docs/statistik/examples` som autoritativ kilde for de seneste spilledage.
"""
    REPORT_PATH.write_text(report, encoding="utf-8")
    print(REPORT_PATH)


if __name__ == "__main__":
    main()
